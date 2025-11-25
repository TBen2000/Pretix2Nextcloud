import requests
import pandas as pd
import os
import logging
from requests.auth import HTTPBasicAuth
import schedule
import time
from datetime import datetime
import pytz
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import base64


DEFAULT_PRETIX_URL = "https://tickets.swdec.de"
DEFAULT_PRETIX_ORGANIZER_SLUG = "kv-stuttgart"
DEFAULT_EXCEL_MAX_COLUMN_WIDTH = 30
DEFAULT_TEMP_DIR_NAME = "jungschartag_automation"
DEFAULT_NEXTCLOUD_URL = "https://jcloud.swdec.de"
DEFAULT_NEXTCLOUD_UPLOAD_DIR = "Jungschartag_Anmeldungen"
DEFAULT_TIMEZONE = "Europe/Berlin"
DEFAULT_INTERVAL_MINUTES = 15
DEFAULT_CHECK_INTERVAL_SECONDS = 60
DEFAULT_RUN_ONCE = False
DEFAULT_LOGGING_LEVEL = "INFO"

logging.basicConfig(level=logging.INFO)

success_on_last_run = False  # track if the last run was successful


class Environment():
    
    def get_pretix_url(self) -> str:
        """
        Return the Pretix base URL from environment variable 'PRETIX_URL'.
        """
        
        pretix_url = self._get_env(name="PRETIX_URL", default=str(DEFAULT_PRETIX_URL))

        if pretix_url.startswith("http://"):
            logging.warning("Environment variable 'PRETIX_URL' starts with 'http://'. It is strongly recommended to use https because sensitive data is transmitted over the internet.")
        elif not pretix_url.startswith("https://"):  # if pretix_url doesn't start with "http://" or "https://"
            pretix_url = f"https://{pretix_url}"
            
        return pretix_url
    
    def get_pretix_api_token(self) -> str:
        """
        Return the Pretix API token from environment variable 'PRETIX_API_TOKEN'.
        """
        
        api_token = self._get_env(name="PRETIX_API_TOKEN", default="[ENV VARIABLE NOT FOUND]")
        
        if api_token and api_token != "[ENV VARIABLE NOT FOUND]":
            return api_token
        
        # get secret instead of env variable:
        
        secret_name = self._get_env(name="PRETIX_API_TOKEN_SECRET_NAME")
        if not secret_name:
            raise ValueError("Environment variable 'PRETIX_API_TOKEN' (or alternatively 'PRETIX_API_TOKEN_SECRET_NAME' for using docker secrets) is not set.")
        
        secret = self._get_secret(secret_name)  
        if not secret:
            raise ValueError(f"Secret {secret_name} for the Pretix API token is empty.")
        
        return secret
    
    def get_pretix_event_slug(self) -> str:
        """
        Return the Pretix event slug from environment variable 'PRETIX_EVENT_SLUG'.
        """
        
        event_slug = self._get_env(name="PRETIX_EVENT_SLUG")
            
        return event_slug
    
    def get_pretix_orgnizer_slug(self) -> str:
        """
        Return the Pretix organizer slug from environment variable 'PRETIX_ORGANIZER_SLUG'.
        """
        
        orgnizer_slug = self._get_env(name="PRETIX_ORGANIZER_SLUG", default=str(DEFAULT_PRETIX_ORGANIZER_SLUG))
            
        return orgnizer_slug
    
    def get_excel_max_column_width(self) -> int:
        """
        Return the max column width for excel files from environment variable 'EXCEL_MAX_COLUMN_WIDTH'.
        """
        
        str_width = self._get_env(name="EXCEL_MAX_COLUMN_WIDTH", default=str(DEFAULT_EXCEL_MAX_COLUMN_WIDTH))

        try:
            width = int(str_width)
        except ValueError:
            logging.error(f"Environment variable 'EXCEL_MAX_COLUMN_WIDTH' must be an integer. Using default value '{DEFAULT_EXCEL_MAX_COLUMN_WIDTH}'.")
            return DEFAULT_EXCEL_MAX_COLUMN_WIDTH

        min_value = 5
        if width < min_value:
            logging.error(f"Environment variable 'EXCEL_MAX_COLUMN_WIDTH' must be at least {min_value}. Using default value '{DEFAULT_EXCEL_MAX_COLUMN_WIDTH}'.")
            return DEFAULT_EXCEL_MAX_COLUMN_WIDTH

        return width
    
    def get_temp_dir_name(self) -> str:
        """
        Return the name of the used subdirectory in the tmp folder from environment variable 'TEMP_DIR_NAME'.
        """
        
        temp_dir_name = self._get_env(name="TEMP_DIR_NAME", default=str(DEFAULT_TEMP_DIR_NAME))

        return temp_dir_name
    
    def get_nextcloud_url(self) -> str:
        """
        Return the Nextcloud URL from environment variable 'NEXTCLOUD_URL'.
        """
        
        pretix_url = self._get_env(name="NEXTCLOUD_URL", default=str(DEFAULT_NEXTCLOUD_URL))

        if pretix_url.startswith("http://"):
            logging.warning("Environment variable 'NEXTCLOUD_URL' starts with 'http://'. It is strongly recommended to use https because sensitive data is transmitted over the internet.")
        elif not pretix_url.startswith("https://"):  # if pretix_url doesn't start with "http://" or "https://"
            pretix_url = f"https://{pretix_url}"
            
        return pretix_url
    
    def get_nextcloud_username(self) -> str:
        """
        Return the Nextcloud username from environment variable 'NEXTCLOUD_USERNAME'.
        """
        
        username = self._get_env(name="NEXTCLOUD_USERNAME", default="[ENV VARIABLE NOT FOUND]")
        
        if username and username != "[ENV VARIABLE NOT FOUND]":
            return username
        
        # get secret instead of env variable:
        
        secret_name = self._get_env(name="NEXTCLOUD_USERNAME_SECRET_NAME")
        if not secret_name:
            raise ValueError("Environment variable 'NEXTCLOUD_USERNAME' (or alternatively 'NEXTCLOUD_USERNAME_SECRET_NAME' for using docker secrets) is not set.")
        
        secret = self._get_secret(secret_name)  
        if not secret:
            raise ValueError(f"Secret {secret_name} for the Nextcloud username is empty.")
        
        return secret
    
    def get_nextcloud_password(self) -> str:
        """
        Return the Nextcloud password from environment variable 'NEXTCLOUD_PASSWORD'.
        """
        
        password = self._get_env(name="NEXTCLOUD_PASSWORD", default="[ENV VARIABLE NOT FOUND]", strip=False)
        
        if password.strip() != "" and password != "[ENV VARIABLE NOT FOUND]":  # if passwort doesn't contain only whitespace
            return password
        
        # get secret instead of env variable:
        
        secret_name = self._get_env(name="NEXTCLOUD_PASSWORD_SECRET_NAME")
        if not secret_name:
            raise ValueError("Environment variable 'NEXTCLOUD_PASSWORD' (or alternatively 'NEXTCLOUD_PASSWORD_SECRET_NAME' for using docker secrets) is not set.")
        
        secret = self._get_secret(secret_name)  
        if not secret:
            raise ValueError(f"Secret {secret_name} for the Nextcloud password is empty.")
        
        return secret
    
    def get_nextcloud_upload_dir(self) -> str:
        """
        Return the Nextcloud upload directory from environment variable 'NEXTCLOUD_UPLOAD_DIR'.
        """
        
        upload_dir = self._get_env(name="NEXTCLOUD_UPLOAD_DIR", default=str(DEFAULT_NEXTCLOUD_UPLOAD_DIR), info_log=True)
        
        upload_dir = upload_dir.strip("/\\")
            
        return upload_dir
    
    def get_timezone(self) -> str:
        """
        Return the timezone from environment variable 'TZ'.
        """
        
        timezone = self._get_env(name="TZ", default=str(DEFAULT_TIMEZONE))
            
        return timezone
    
    def get_interval_minutes(self) -> int:
        """
        Return the interval in minutes to run the main function in loop from environment variable 'INTERVAL_MINUTES'.
        """
        
        str_minutes = self._get_env(name="INTERVAL_MINUTES", default=str(DEFAULT_INTERVAL_MINUTES))

        try:
            minutes = int(str_minutes)
        except ValueError:
            logging.error(f"Environment variable 'INTERVAL_MINUTES' must be an integer. Using default value '{DEFAULT_INTERVAL_MINUTES}'.")
            return DEFAULT_INTERVAL_MINUTES

        min_value = 1
        if minutes < min_value:
            logging.error(f"Environment variable 'INTERVAL_MINUTES' must be at least {min_value}. Using default value '{DEFAULT_INTERVAL_MINUTES}'.")
            return DEFAULT_INTERVAL_MINUTES

        return minutes
    
    def get_check_interval_seconds(self) -> int:
        """
        Return the interval in second to check if the time to wait for running again is over. From environment variable 'CHECK_INTERVAL_SECONDS'.
        """
        
        str_seconds = self._get_env(name="CHECK_INTERVAL_SECONDS", default=str(DEFAULT_CHECK_INTERVAL_SECONDS))

        try:
            seconds = int(str_seconds)
        except ValueError:
            logging.error(f"Environment variable 'CHECK_INTERVAL_SECONDS' must be an integer. Using default value '{DEFAULT_CHECK_INTERVAL_SECONDS}'.")
            return DEFAULT_CHECK_INTERVAL_SECONDS

        min_value = 1
        if seconds < min_value:
            logging.error(f"Environment variable 'CHECK_INTERVAL_SECONDS' must be at least {min_value}. Using default value '{DEFAULT_CHECK_INTERVAL_SECONDS}'.")
            return DEFAULT_CHECK_INTERVAL_SECONDS

        return seconds
   
    def get_run_once(self) -> bool:
        """
        Return wheather to run once from environment variable 'RUN_ONCE'.
        """
        
        run_once = self._get_env(name="RUN_ONCE", default=str(DEFAULT_RUN_ONCE))
        run_once_lower = run_once.lower()
        
        if run_once_lower == "true":
            return True
        if run_once_lower == "false":
            return False
        
        raise ValueError(f"Environment variable 'RUN_ONCE' must be either 'true' or 'false'. Current value: '{run_once}'")

    def get_logging_level(self) -> int:
        """
        Return logging level from environment variable 'LOGGING_LEVEL'.
        """
        logging_level = self._get_env(name="LOGGING_LEVEL", default=str(DEFAULT_LOGGING_LEVEL))
        logging_level = logging_level.lower()

        levels = {
            "debug": logging.DEBUG,
            "info": logging.INFO,
            "warning": logging.WARNING,
            "error": logging.ERROR,
        }

        if logging_level in levels:
            return levels[logging_level]

        raise ValueError(f"Environment variable 'LOGGING_LEVEL' must be either 'debug', 'info', 'warning', or 'error'. Current value: '{logging_level}'.")
    
    
    def _get_env(self, name: str, default: str = "", strip: bool = True, info_log: bool = False) -> str:
        """
        Get environment variable with optional default value.

        Logs a warning when default is used and an error when a required
        environment variable is missing.
        """

        env = os.getenv(name, "")
        
        if strip is True:
            env = env.strip()

        if not env:
            
            default = default.strip()
            if default:
                env = default
                
                logging_text = f"Environment Variable '{name}' is not set. Using default '{default}'."
                if info_log is True:
                    logging.info(logging_text)
                else:
                    logging.debug(logging_text)
                
            else:
                raise ValueError(f"Environment Variable '{name}' is not set.")

        env = self._decode_if_base64(env, prefix="BASE64:")

        return env


    def _get_secret(self, name: str) -> str:
        """
        Read a docker-style secret from a file and return its contents. Handles base64-encoded secrets if prefixed with "BASE64:".

        Returns an empty string and logs an error on failures.
        """
        
        path = os.path.join("/run/secrets", name)

        try:
            with open(path, "r") as f:
                secret = f.read().strip("\n")

        except Exception as e:
            logging.error(f"Error reading secret from {path}: {e}")
            secret = ""

        secret = self._decode_if_base64(secret, prefix="BASE64:")

        return secret


    def _decode_if_base64(self, data: str, prefix: str) -> str:
        """
        Decode the given data from base64 if it starts with the specified prefix.
        """

        if isinstance(data, str) and data.startswith(prefix):
            try:
                b64_content = data[len(prefix) :]
                data = base64.b64decode(b64_content).decode("utf-8").strip("\n")
            except Exception as e:
                logging.error(f"Error decoding base64: {e}")
                data = ""

        return data


class Dataframe:
    last_raw_df = pd.DataFrame()  # global variable to store last fetched raw dataframe

    def __init__(self, success_on_last_run: bool = False):
        """
        Initialize, fetch data from Pretix API and load into different desired dataframes.
        """
        env = Environment()

        pretix_url = env.get_pretix_url()
        pretix_organizer = env.get_pretix_orgnizer_slug()
        pretix_event = env.get_pretix_event_slug()
        pretix_api_token = env.get_pretix_api_token()

        self.time_zone = env.get_timezone()

        self.pretix_api_url = os.path.join(pretix_url, "api/v1/organizers", pretix_organizer, "events", pretix_event)
        self.headers = {"Authorization": f"Token {pretix_api_token}"}
        self.raw_df = self._get_raw_df()

        if success_on_last_run and self.__class__.last_raw_df.equals(self.raw_df):
            raise Exception("No changes in data since last fetch.")
        self.__class__.last_raw_df = self.raw_df

        self.sorted_df = self._get_sorted_df()

        self.towns_list = self._get_question_choices_by_text("Ich melde mich über folgende Ortschaft an")
        self.town_dfs = self._get_town_dfs()
        self.numbers_overview = self._get_numbers_df()


    def _get_questions(self) -> dict:
        """
        Fetch all questions from Pretix API and return a mapping of question ID to question text.
        """

        url = f"{self.pretix_api_url}/questions/"
        questions = {}

        while url:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            data = response.json()

            for q in data["results"]:
                question_text = q["question"].get("de") or next(iter(q["question"].values()))
                questions[q["id"]] = question_text

            url = data["next"]  # Pagination

        return questions

    def _get_question_choices_by_text(self, question_str: str) -> list:
        """
        Given the clear-text question name (question_str), find all question IDs that
        have that visible text and fetch their possible answer options from the
        Pretix API. If multiple question IDs match the same question_str, the
        returned answer options are merged uniquely and this fact is logged.

        Returns a sorted list of unique answer option strings.
        """

        if not question_str.strip():
            logging.error("Empty question text provided to get_question_choices_by_text")
            return []

        # build a fresh mapping of id -> text (ensures up-to-date data)
        question_map = self._get_questions()

        # case-insensitive match on the visible question text
        target = question_str.strip().lower()
        matching_qids = [qid for qid, text in question_map.items() if (text or "").strip().lower() == target]

        if not matching_qids:
            logging.warning(f"No question IDs found for question text '{question_str}'")
            return []

        all_choices = set()

        for qid in matching_qids:
            url = f"{self.pretix_api_url}/questions/{qid}/"
            try:
                resp = requests.get(url, headers=self.headers)
                resp.raise_for_status()
                data = resp.json()

                # pretix may expose options under different keys depending on API version/implementation
                options = []
                for key in ("options", "choices", "answers", "options_list", "question_options"):
                    if key in data and isinstance(data[key], list):
                        options = data[key]
                        break

                # fallback: sometimes the question detail may include nested structures
                if not options:
                    # try to find any list-valued field in response that looks like options
                    for v in data.values():
                        if isinstance(v, list) and v and isinstance(v[0], (str, dict)):
                            options = v
                            break

                # extract a human-readable string from each option entry.
                def _extract_choice_text(opt) -> str | None:
                    # plain string option
                    if isinstance(opt, str):
                        return opt.strip() or None

                    if not isinstance(opt, dict):
                        return None

                    # prefer common direct string keys
                    for k in ("label", "text", "answer", "name", "title", "display"):
                        v = opt.get(k)
                        if isinstance(v, str) and v.strip():
                            return v.strip()

                        # if the value is a translations dict (e.g. {"de": "..."}), prefer German then English
                        if isinstance(v, dict):
                            for lang in ("de", "de-DE", "en", "en-US"):
                                if (
                                    lang in v
                                    and isinstance(v[lang], str)
                                    and v[lang].strip()
                                ):
                                    return v[lang].strip()
                            # fallback to first available string in translations
                            for vv in v.values():
                                if isinstance(vv, str) and vv.strip():
                                    return vv.strip()

                    # if no labelled text found, try to find any string value in the dict
                    for vv in opt.values():
                        if isinstance(vv, str) and vv.strip():
                            return vv.strip()

                    # nothing human-readable found
                    return None

                for opt in options:
                    text_val = _extract_choice_text(opt)
                    if text_val:
                        all_choices.add(text_val)

            except Exception as e:
                logging.error(f"Error fetching choices for question id {qid}: {e}")

        if len(matching_qids) > 1:
            logging.info(f"Multiple question IDs {matching_qids} map to the same question text '{question_str}'. Merged unique choices ({len(all_choices)} unique).")

        return sorted(all_choices)

    def _get_items(self) -> dict:
        """
        Fetch all items from Pretix API and return a mapping of item ID to item name.
        """

        url = f"{self.pretix_api_url}/items/"
        items = {}

        while url:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            data = response.json()

            for i in data["results"]:
                item_name = i["name"].get("de") or next(iter(i["name"].values()))
                items[i["id"]] = item_name

            url = data["next"]  # Pagination

        return items

    def _get_orders(self) -> list:
        """
        Fetch all orders from Pretix API and return as a list of order dicts.
        """

        url = f"{self.pretix_api_url}/orders/"
        orders = []

        while url:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            data = response.json()

            orders.extend(data["results"])
            url = data["next"]  # Pagination

        return orders

    def _get_unique_column_name(self, base_name: str, existing_columns: list) -> str:
        """
        Generate a unique column name by adding (#2), (#3), etc. suffix if needed.

        Args:
            base_name: The desired column name
            existing_columns: List of column names that already exist

        Returns:
            Unique column name (either base_name or base_name with " (#n)" suffix)
        """

        if base_name not in existing_columns:
            return base_name

        # Find the next available number
        counter = 2
        while f"{base_name} (#{counter})" in existing_columns:
            counter += 1

        return f"{base_name} (#{counter})"

    def _orders_to_dataframe(self, orders: list, question_map: dict, item_map: dict) -> pd.DataFrame:
        """
        Convert orders to a pandas DataFrame, resolving question and item names.
        """

        rows = []
        # Track which question texts we've already seen to handle duplicate question texts
        question_text_mapping = {}  # maps original qtext -> unique column name

        for order in orders:
            invoice = order.get("invoice_address", {}) or {}
            order_info = {
                "order_code": order["code"],
                "status": order["status"],
                "email": order["email"],
                "total": order["total"],
                "date": order["datetime"],
                "invoice_name": invoice.get("name", ""),
                "invoice_company": invoice.get("company", ""),
                "invoice_street": invoice.get("street", ""),
                "invoice_zipcode": invoice.get("zipcode", ""),
                "invoice_city": invoice.get("city", ""),
                "invoice_country": invoice.get("country", ""),
                "invoice_vat_id": invoice.get("vat_id", ""),
            }

            for position in order["positions"]:
                # resolve item id
                item_id = position.get("item")
                if isinstance(item_id, dict):
                    item_name = item_id["name"].get("de") or next(
                        iter(item_id["name"].values())
                    )
                else:
                    item_name = item_map.get(item_id, f"Item {item_id}")

                # attendee name from attendee_name_parts
                name_parts = position.get("attendee_name_parts", {}) or {}
                attendee_firstname = name_parts.get("given_name", "")
                attendee_lastname = name_parts.get("family_name", "")

                pos_info = {
                    "position_id": position["id"],
                    "item_id": item_id
                    if not isinstance(item_id, dict)
                    else item_id["id"],
                    "item_name": item_name,
                    "price": position["price"],
                    "attendee_firstname": attendee_firstname,
                    "attendee_lastname": attendee_lastname
                }

                # answers for questions
                questions = {}
                for answer in position.get("answers", []):
                    qid = answer.get("question")
                    answer_text = answer.get("answer", "")

                    if isinstance(qid, dict):
                        qid = qid["id"]

                    qtext = question_map.get(qid, f"question_{qid}")

                    # Generate unique column name for this question text
                    if qtext not in question_text_mapping:
                        question_text_mapping[qtext] = self._get_unique_column_name(qtext, list(question_text_mapping.values()))

                    unique_col_name = question_text_mapping[qtext]
                    questions[unique_col_name] = answer_text

                row = {**order_info, **pos_info, **questions}
                rows.append(row)

        df = pd.DataFrame(rows)

        # Ensure all questions from Pretix are present as columns even if
        # no attendee has answered them yet. Handle duplicate question texts with (#2), (#3), etc.
        for qid, qtext in question_map.items():
            # Generate unique column name for this question text
            if qtext not in question_text_mapping:
                unique_col_name = self._get_unique_column_name(
                    qtext, list(question_text_mapping.values()) + list(df.columns)
                )
                question_text_mapping[qtext] = unique_col_name
            else:
                unique_col_name = question_text_mapping[qtext]

            # Add empty column if it doesn't exist yet
            if unique_col_name not in df.columns:
                df[unique_col_name] = ""

        return df

    def _get_raw_df(self) -> pd.DataFrame:
        """
        Fetch raw data from Pretix API and return as a pandas DataFrame.
        """

        question_map = self._get_questions()
        item_map = self._get_items()
        orders = self._get_orders()
        df = self._orders_to_dataframe(orders, question_map, item_map)

        logging.info("Fetched raw data from Pretix API.")

        return df

    def _get_sorted_df(self) -> pd.DataFrame:
        """
        Process raw dataframe to create a sorted dataframe with required columns.
        """

        df = self.raw_df

        # removed all cancelled registrations
        df = df[df["status"] != "c"]

        # rename needed columns
        renames = {
            "order_code": "Bestellnummer",
            "email": "E-Mail",
            "date": "Anmeldedatum",
            "item_name": "Art",
            "attendee_firstname": "Vorname",
            "attendee_lastname": "Nachname",
            "Essensunverträglichkeiten": "Essensunverträglichkeiten Ja/Nein",
            "Welche Unverträglichkeiten?": "Essensunverträglichkeiten",
            "Ich melde mich über folgende Ortschaft an": "Ortschaft",
            "Ich biete eine Fahrgemeinschaft an": "Fahrer Angebot",
            "Ich bin Ortsverantwortlicher.": "Ortsverantwortlicher",
            "Telefonnummer": "Telefonnummer Mitarbeiter"
        }
        df = df.rename(columns=renames)

        # combine "Telefonnummer der Eltern" and "Telefonnummer Mitarbeiter" to one column "Telefonnummer"
        df["Telefonnummer"] = df["Telefonnummer der Eltern"].combine_first(df["Telefonnummer Mitarbeiter"])

        # filter for columns and set their order
        wanted_columns = [
            "Ortschaft",
            "Art",
            "Nachname",
            "Vorname",
            "Telefonnummer",
            "E-Mail",
            "Ernährung",
            "Essensunverträglichkeiten",
            "Sonstiges",
            "Ortsverantwortlicher",
            "Fahrer Angebot",
            "Anmeldedatum",
            "Bestellnummer"
        ]
        df = df.filter(wanted_columns)

        # change date format
        df["Anmeldedatum"] = (pd.to_datetime(df["Anmeldedatum"], utc=True).dt.tz_convert(self.time_zone).dt.strftime("%Y-%m-%d %H:%M"))

        # sort (by "Ortschaft", then by "Art", then by "Nachname" and then by "Vorname") and reset index numbers
        df = df.sort_values(by=["Ortschaft", "Art", "Nachname", "Vorname"], ascending=True)
        df.index = range(1, len(df) + 1)

        logging.info("Sorted raw data.")

        return df

    def _get_town_dfs(self) -> dict[str, pd.DataFrame]:
        """
        Process sorted dataframe to create a dictionary of dataframes filtered by town.
        """

        df = self.sorted_df

        # filter for columns and set their order
        wanted_columns = [
            "Ortschaft",
            "Art",
            "Nachname",
            "Vorname",
            "Telefonnummer",
            "E-Mail",
            "Ernährung",
            "Essensunverträglichkeiten",
            "Sonstiges",
            "Fahrer Angebot",
            "Anmeldedatum"
        ]
        df = df.filter(wanted_columns)

        # sort by town:
        df_by_town_dict = {}
        for town in self.towns_list:
            # filter by town, drop column "Ortschaft" and reset index numbers
            town_df = df[df["Ortschaft"] == town]
            town_df = town_df.drop(columns=["Ortschaft"])
            town_df.index = range(1, len(town_df) + 1)

            df_by_town_dict[town] = town_df

        logging.info("Filtered sorted data by town.")

        return df_by_town_dict

    def _get_numbers_df(self) -> pd.DataFrame:
        """
        Calculate and return a dataframe with counts of "Jungscharler", "Mitarbeiter", and total by town.
        """

        numbers_df = pd.DataFrame(
            {"Ortschaft": [], "Jungscharler": [], "Mitarbeiter": [], "Gesamt": []}
        )

        # make Ortschaft the index
        numbers_df.set_index("Ortschaft", inplace=True)

        df = self.sorted_df

        number_of_kids = len(df[df["Art"].str.contains("Jungscharler", na=False)])
        number_of_staff = len(df[df["Art"].str.contains("Mitarbeiter", na=False)])
        number_total = number_of_kids + number_of_staff

        # add row to numbers_df
        numbers_df.loc["GESAMT"] = [number_of_kids, number_of_staff, number_total]

        # filter by town:
        for town in self.towns_list:
            town_df = df[df["Ortschaft"] == town]

            town_kids = len(
                town_df[town_df["Art"].str.contains("Jungscharler", na=False)]
            )
            town_staff = len(
                town_df[town_df["Art"].str.contains("Mitarbeiter", na=False)]
            )
            town_total = town_kids + town_staff

            # add row to numbers_df
            numbers_df.loc[town] = [town_kids, town_staff, town_total]

        return numbers_df


class Excel:
    def __init__(self):
        """
        Initialize the Excel helper class, setting up a temporary directory for Excel files.
        """
        env = Environment()
        self.max_column_width = env.get_excel_max_column_width()
        temp_dir_name = env.get_temp_dir_name()
        
        self.temp_dir = os.path.join(tempfile.gettempdir(), temp_dir_name)
        Path(self.temp_dir).mkdir(parents=True, exist_ok=True)
        
    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize the filename by replacing or removing invalid characters.
        """
        filename = filename.strip()
        filename = filename.replace('\\n', ' ')
        filename = filename.replace('\\r', ' ')
        filename = filename.replace('\\t', ' ')
        filename = filename.replace('<', '_')
        filename = filename.replace('>', '_')
        filename = filename.replace(':', '')
        filename = filename.replace('"', '')
        filename = filename.replace('/', '+')
        filename = filename.replace('\\', '_')
        filename = filename.replace('|', '_')
        filename = filename.replace('?', '')
        filename = filename.replace('*', '')
    
        return filename

    def save_to_excel(self, df: pd.DataFrame, filename: str) -> str:
        """
        Save the given DataFrame to an Excel file with the specified filename in the temporary directory.
        Returns the path to the saved Excel file.
        """
        
        filename = self._sanitize_filename(filename)

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        sheet_name = filename.removesuffix(".xlsx")

        path = os.path.join(self.temp_dir, filename)

        # write dataframe to excel
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=True, freeze_panes=(1, 1))
            worksheet = writer.sheets[sheet_name]

            # adjust column widths including index column and header
            for idx, col in enumerate(df.columns, start=2):  # start=2 because index is in column 1
                max_length = max(
                    len(str(col)),  # header length
                    df[col].astype(str).map(len).max()  # data length
                )
                length = min(max_length + 2, self.max_column_width)  # cap the width via constant
                
                worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = length

            # adjust index column width
            max_index_length = max(
                len(str(df.index.name or "")),  # index name length
                df.index.astype(str).map(len).max()  # index data length
            )
            worksheet.column_dimensions["A"].width = max_index_length + 2

        logging.info(f"Created Excel file '{path}'.")

        return path

    def add_filter(self, path_to_excel_file: str) -> str:
        """
        Add filtering function to an existing Excel file.
        """

        path = path_to_excel_file

        if not os.path.isfile(path):
            logging.warning(f"Cannot add filter to '{path}': No such file.")
            return path
        
        elif not path.endswith(".xlsx"):
            logging.warning(f"Cannot add filter to '{path}': File is not an excel file.")
            return path

        wb = load_workbook(path)
        ws = wb.active

        # Determine the range of the data
        min_col = 2  # assuming index is in column A
        min_row = 1
        max_col = ws.max_column
        max_row = ws.max_row
        table_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

        # Create a table
        tab = Table(displayName="Table1", ref=table_range)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(tab)

        wb.save(path)

        logging.info(f"Added filter to Excel file '{path}'.")

        return path

    def delete_excel(self, path_to_excel_file: str) -> None:
        """
        Delete a temporary Excel file.
        """
        try:
            if os.path.isfile(path_to_excel_file):
                os.remove(path_to_excel_file)
                logging.info(f"Deleted temporary Excel file '{path_to_excel_file}'.")
        except Exception as e:
            logging.error(f"Error deleting file '{path_to_excel_file}': {e}")


class Nextcloud:
    def __init__(self):
        """
        Initialize Nextcloud connection with environment variables.
        """
        env = Environment()
        
        nextcloud_url = env.get_nextcloud_url()
        self.username = env.get_nextcloud_username()
        self.password = env.get_nextcloud_password()
        self.upload_dir = env.get_nextcloud_upload_dir()
        self.time_zone = env.get_timezone()
        
        self.base_url = os.path.join(nextcloud_url, "remote.php/dav/files", self.username)
        
        self.upload_dir_url = os.path.join(self.base_url, self.upload_dir)

    
    def _get_parent_directories(self, path: str) -> list[str]:
        """
        Get all parent directories of a given path.
        e.g. for path "A/B/C" it returns ["A", "A/B", "A/B/C"]
        """
        
        # ensure that the path does not end with a slash
        path = path.rstrip(os.sep)
        
        parent_dirs = []
        
        # walk through the path and build the parent directories
        while path != os.path.dirname(path):
            parent_dirs.append(path)
            path = os.path.dirname(path)

        # reverse the list to have the directories in hierarchical order from top to bottom.
        return parent_dirs[::-1]

    def create_upload_directory(self) -> None:
        """
        Create the upload directory on Nextcloud if it does not exist.
        """
        
        # check if directory exists:
        r = requests.request(
            method="PROPFIND",
            url=self.upload_dir_url,
            auth=HTTPBasicAuth(self.username, self.password)
        )
        
        if r.status_code == 207:
            logging.info(f"Upload directory already exists ('{self.upload_dir_url}')")
            return
        
        # ceate directory
        try:
            for dir in self._get_parent_directories(self.upload_dir):
                
                r = requests.request(
                    method="MKCOL",
                    url=os.path.join(self.base_url, dir),
                    auth=HTTPBasicAuth(self.username, self.password)
                )
                
                if r.status_code not in [201, 405]:
                    logging.error(f"Error creating upload directory: {r.status_code} - {r.text}")
                    return
        
            logging.info(f"Upload directory created ('{self.upload_dir_url}')")
            
        except Exception as e:
            logging.error(f"Error creating upload directory: {e}")

    def _upload_file(self, filename: str, data: bytes) -> None:
        """
        Upload a file to Nextcloud via WebDAV.
        """

        try:
            r = requests.put(
                url=os.path.join(self.upload_dir_url, filename),
                data=data,
                auth=HTTPBasicAuth(self.username, self.password)
            )
            
            if r.status_code in (200, 201, 204):
                logging.info(f"File '{filename}' uploaded successfully.")
            else:
                logging.error(f"Error uploading file '{filename}': {r.status_code} - {r.text}")
                
        except Exception as e:
            logging.error(f"Network error uploading file '{filename}': {e}")

    def upload_excel(self, source_file: str) -> None:
        """
        Upload an Excel file to Nextcloud.
        """

        try:
            filename = os.path.basename(source_file)
            with open(source_file, "rb") as f:
                data = f.read()
            self._upload_file(filename, data)
        except Exception as e:
            logging.error(f"Error uploading Excel file '{source_file}': {e}")

    def upload_last_updated(self) -> None:
        """
        Upload a timestamp file indicating the last update time.
        """

        filename = "Last_Updated.txt"

        tz = pytz.timezone(self.time_zone)
        data = "Last updated:\n" + datetime.now(tz=tz).strftime("%d.%m.%Y %H:%M")
        

        self._upload_file(filename, data.encode("utf-8"))


def main():
    """
    Main function to generate Excel files and upload them to Nextcloud.
    """

    global success_on_last_run

    try:
        dataframe = Dataframe(success_on_last_run)
        excel = Excel()
        nc = Nextcloud()

        nc.create_upload_directory()

        # generate, upload and delete raw data file
        filepath = excel.save_to_excel(dataframe.raw_df, "Raw_Data")
        nc.upload_excel(filepath)
        excel.delete_excel(filepath)

        # generate, upload and delete all attendees file
        filepath = excel.save_to_excel(dataframe.sorted_df, "Alle")
        filepath = excel.add_filter(filepath)
        nc.upload_excel(filepath)
        excel.delete_excel(filepath)

        # generate, upload and delete town-wise attendees files
        for town in dataframe.town_dfs:
            df = dataframe.town_dfs[town]

            filepath = excel.save_to_excel(df, town)
            nc.upload_excel(filepath)
            excel.delete_excel(filepath)

        # generate, upload and delete all numbers_overview file
        filepath = excel.save_to_excel(dataframe.numbers_overview, "Anmeldezahlen")
        nc.upload_excel(filepath)
        excel.delete_excel(filepath)

        nc.upload_last_updated()

        success_on_last_run = True

    except Exception as e:
        if str(e) == "No changes in data since last fetch.":
            logging.info("No changes detected since last fetch. Skipping upload.")
            try:
                nc = Nextcloud()
                nc.upload_last_updated()
            except Exception as e:
                logging.error(f"An error occurred while uploading last updated timestamp: {e}")

        else:
            logging.error(f"An error occurred during execution: {e}")
            success_on_last_run = False


if __name__ == "__main__":
    env = Environment()
    level = env.get_logging_level()
    
    # remove all handlers associated with the root logger object and set desired logging level
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)
    logging.basicConfig(level=level)
    
    # run main immediately on startup
    main()

    if env.get_run_once() is True:
        logging.info("RUN_ONCE is set to true. Exiting now after single run.")
        exit(0)

    # schedule main to run at configured intervals
    schedule.every(env.get_interval_minutes()).minutes.do(main)

    # keep the scheduler running
    while True:
        schedule.run_pending()
        time.sleep(env.get_check_interval_seconds())  # check periodically if a task needs to run
