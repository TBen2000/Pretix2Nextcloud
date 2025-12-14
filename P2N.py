# General Pretix2Nextcloud helper classes
# Needs to be imported in specific instances of Pretix2Nextcloud (e.g. in kv-stuttgart-jungschartag.py)

import requests
import pandas as pd
import os
import logging
from requests.auth import HTTPBasicAuth
import schedule
import time
from datetime import datetime
import pytz
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import base64
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

logging.basicConfig(level=logging.INFO)


class Environment:

    def set_defaults(
        self,
        default_pretix_url: str = None,
        default_pretix_organizer_slug: str = None,
        default_excel_max_column_width: int = None,
        default_temp_dir_name: str = None,
        default_nextcloud_url: str = None,
        default_nextcloud_upload_dir: str = None,
        default_timezone: str = None,
        default_interval_minutes: int = None,
        default_check_interval_seconds: int = None,
        default_run_once: str = None,
        default_logging_level: str = None,
    ):
        """
        Sets default values for the class as global class variables. Only parameters that are not None will be applied.
        """

        # Iterate over all parameters and assign them to the class if they are not None
        for attr, value in locals().items():
            if attr == "self":
                continue  # skip the instance reference
            if value is not None:
                setattr(self.__class__, attr, value)
                
    def _get_class_variable_value(self, name: str):
        """
        Returns value for a class variable set in set_defaults().
        """
        
        value = getattr(self.__class__, name, None)
        if value is None:
            return ""
        return value

    def get_pretix_url(self) -> str:
        """
        Return the Pretix base URL from environment variable 'PRETIX_URL'.
        """
        
        try:
            default = str(self._get_class_variable_value("default_pretix_url"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_pretix_url can't be stringified. Check the value you entered while calling set_defaults() function."
            )

        pretix_url = self._get_env(name="PRETIX_URL", default=default)

        if pretix_url.startswith("http://"):
            logging.warning(
                "Environment variable 'PRETIX_URL' starts with 'http://'. It is strongly recommended to use https because sensitive data is transmitted over the internet."
            )
        elif not pretix_url.startswith(
            "https://"
        ):  # if pretix_url doesn't start with "http://" or "https://"
            pretix_url = f"https://{pretix_url}"

        return pretix_url

    def get_pretix_api_token(self) -> str:
        """
        Return the Pretix API token from environment variable 'PRETIX_API_TOKEN'.
        Alternatively: Read name of Docker secret from 'PRETIX_API_TOKEN_SECRET_NAME' and return secret.
        """

        try:
            api_token = self._get_env(name="PRETIX_API_TOKEN")
            return api_token
        except ValueError:
            logging.debug(
                "Environment variable 'PRETIX_API_TOKEN' is not set. Trying for Docker secret."
            )

        # get secret instead of env variable:
        try:
            secret_name = self._get_env(name="PRETIX_API_TOKEN_SECRET_NAME")
        except ValueError:
            raise ValueError(
                "Environment variable 'PRETIX_API_TOKEN' (or alternatively 'PRETIX_API_TOKEN_SECRET_NAME' for using docker secrets) is not set."
            )

        secret = self._get_secret(secret_name)
        if not secret:
            raise ValueError(f"Secret {secret_name} for the Pretix API token is empty.")

        return secret

    def get_pretix_event_slug(self) -> str:
        """
        Return the Pretix event slug from environment variable 'PRETIX_EVENT_SLUG'.
        """

        event_slug = self._get_env(name="PRETIX_EVENT_SLUG")

        return event_slug

    def get_pretix_orgnizer_slug(self) -> str:
        """
        Return the Pretix organizer slug from environment variable 'PRETIX_ORGANIZER_SLUG'.
        """

        try:
            default = str(self._get_class_variable_value("default_pretix_organizer_slug"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_pretix_organizer_slug can't be stringified. Check the value you entered while calling set_defaults() function."
            )

        orgnizer_slug = self._get_env(name="PRETIX_ORGANIZER_SLUG", default=default)

        return orgnizer_slug

    def get_excel_max_column_width(self) -> int:
        """
        Return the max column width for excel files from environment variable 'EXCEL_MAX_COLUMN_WIDTH'.
        """

        try:
            default = int(self._get_class_variable_value("default_excel_max_column_width"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_excel_max_column_width can't be translated to integer. Check the value you entered while calling set_defaults() function."
            )

        str_width = self._get_env(name="EXCEL_MAX_COLUMN_WIDTH", default=str(default))

        try:
            width = int(str_width)
        except ValueError:
            logging.error(
                f"Environment variable 'EXCEL_MAX_COLUMN_WIDTH' must be an integer. Using default value '{default}'."
            )
            return default

        min_value = 5
        if width < min_value:
            logging.error(
                f"Environment variable 'EXCEL_MAX_COLUMN_WIDTH' must be at least {min_value}. Using default value '{default}'."
            )
            return default

        return width

    def get_temp_dir_name(self) -> str:
        """
        Return the name of the used subdirectory in the tmp folder from environment variable 'TEMP_DIR_NAME'.
        """

        try:
            default = str(self._get_class_variable_value("default_temp_dir_name"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_temp_dir_name can't be stringified. Check the value you entered while calling set_defaults() function."
            )

        temp_dir_name = self._get_env(name="TEMP_DIR_NAME", default=default)

        return temp_dir_name

    def get_nextcloud_url(self) -> str:
        """
        Return the Nextcloud URL from environment variable 'NEXTCLOUD_URL'.
        """

        try:
            default = str(self._get_class_variable_value("default_nextcloud_url"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_nextcloud_url can't be stringified. Check the value you entered while calling set_defaults() function."
            )

        pretix_url = self._get_env(
            name="NEXTCLOUD_URL", default=default
        )

        if pretix_url.startswith("http://"):
            logging.warning(
                "Environment variable 'NEXTCLOUD_URL' starts with 'http://'. It is strongly recommended to use https because sensitive data is transmitted over the internet."
            )
        elif not pretix_url.startswith(
            "https://"
        ):  # if pretix_url doesn't start with "http://" or "https://"
            pretix_url = f"https://{pretix_url}"

        return pretix_url

    def get_nextcloud_username(self) -> str:
        """
        Return the Nextcloud username from environment variable 'NEXTCLOUD_USERNAME'.
        Alternatively: Read name of Docker secret from 'NEXTCLOUD_USERNAME_SECRET_NAME' and return secret.
        """

        try:
            username = self._get_env(name="NEXTCLOUD_USERNAME")
            return username
        except ValueError:
            logging.debug(
                "Environment variable 'NEXTCLOUD_USERNAME' is not set. Trying for Docker secret."
            )

        # get secret instead of env variable:

        try:
            secret_name = self._get_env(name="NEXTCLOUD_USERNAME_SECRET_NAME")
        except ValueError:
            raise ValueError(
                "Environment variable 'NEXTCLOUD_USERNAME' (or alternatively 'NEXTCLOUD_USERNAME_SECRET_NAME' for using docker secrets) is not set."
            )

        secret = self._get_secret(secret_name)
        if not secret:
            raise ValueError(
                f"Secret {secret_name} for the Nextcloud username is empty."
            )

        return secret

    def get_nextcloud_password(self) -> str:
        """
        Return the Nextcloud password from environment variable 'NEXTCLOUD_PASSWORD'.
        Alternatively: Read name of Docker secret from 'NEXTCLOUD_PASSWORD_SECRET_NAME' and return secret.
        """

        try:
            password = self._get_env(name="NEXTCLOUD_PASSWORD", strip=False)
            return password
        except ValueError:
            logging.debug(
                "Environment variable 'NEXTCLOUD_PASSWORD' is not set. Trying for Docker secret."
            )

        # get secret instead of env variable:
        try:
            secret_name = self._get_env(name="NEXTCLOUD_PASSWORD_SECRET_NAME")
        except ValueError:
            raise ValueError(
                "Environment variable 'NEXTCLOUD_PASSWORD' (or alternatively 'NEXTCLOUD_PASSWORD_SECRET_NAME' for using docker secrets) is not set."
            )

        secret = self._get_secret(secret_name)
        if not secret:
            raise ValueError(
                f"Secret {secret_name} for the Nextcloud password is empty."
            )

        return secret

    def get_nextcloud_upload_dir(self) -> str:
        """
        Return the Nextcloud upload directory from environment variable 'NEXTCLOUD_UPLOAD_DIR'.
        """
        
        try:
            default = str(self._get_class_variable_value("default_nextcloud_upload_dir"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_nextcloud_upload_dir can't be stringified. Check the value you entered while calling set_defaults() function."
            )

        upload_dir = self._get_env(
            name="NEXTCLOUD_UPLOAD_DIR",
            default=default,
            info_log=True,
        )

        upload_dir = upload_dir.strip("/\\")

        return upload_dir

    def get_timezone(self) -> str:
        """
        Return the timezone from environment variable 'TZ'.
        """

        try:
            default = str(self._get_class_variable_value("default_timezone"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_timezone can't be stringified. Check the value you entered while calling set_defaults() function."
            )
            
        timezone = self._get_env(name="TZ", default=default)

        return timezone

    def get_interval_minutes(self) -> int:
        """
        Return the interval in minutes to run the main function in loop from environment variable 'INTERVAL_MINUTES'.
        """

        try:
            default = int(self._get_class_variable_value("default_interval_minutes"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_interval_minutes can't be translated to integer. Check the value you entered while calling set_defaults() function."
            )
            
        str_minutes = self._get_env(
            name="INTERVAL_MINUTES", default=str(default)
        )

        try:
            minutes = int(str_minutes)
        except ValueError:
            logging.error(
                f"Environment variable 'INTERVAL_MINUTES' must be an integer. Using default value '{default}'."
            )
            return default

        min_value = 1
        if minutes < min_value:
            logging.error(
                f"Environment variable 'INTERVAL_MINUTES' must be at least {min_value}. Using default value '{default}'."
            )
            return default

        return minutes

    def get_check_interval_seconds(self) -> int:
        """
        Return the interval in second to check if the time to wait for running again is over. From environment variable 'CHECK_INTERVAL_SECONDS'.
        """

        try:
            default = int(self._get_class_variable_value("default_check_interval_seconds"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_check_interval_seconds can't be translated to integer. Check the value you entered while calling set_defaults() function."
            )
        
        str_seconds = self._get_env(
            name="CHECK_INTERVAL_SECONDS", default=str(default)
        )

        try:
            seconds = int(str_seconds)
        except ValueError:
            logging.error(
                f"Environment variable 'CHECK_INTERVAL_SECONDS' must be an integer. Using default value '{default}'."
            )
            return default

        min_value = 1
        if seconds < min_value:
            logging.error(
                f"Environment variable 'CHECK_INTERVAL_SECONDS' must be at least {min_value}. Using default value '{default}'."
            )
            return default

        return seconds

    def get_run_once(self) -> bool:
        """
        Return wheather to run once from environment variable 'RUN_ONCE'.
        """

        try:
            default = str(self._get_class_variable_value("default_run_once"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_run_once can't be stringified. Check the value you entered while calling set_defaults() function."
            )
            
        run_once = self._get_env(name="RUN_ONCE", default=default)
        run_once_lower = run_once.lower()

        if run_once_lower == "true":
            return True
        if run_once_lower == "false":
            return False

        raise ValueError(
            f"Environment variable 'RUN_ONCE' must be either 'true' or 'false'. Current value: '{run_once}'"
        )

    def get_logging_level(self) -> int:
        """
        Return logging level from environment variable 'LOGGING_LEVEL'.
        """

        try:
            default = str(self._get_class_variable_value("default_logging_level"))
        except Exception:
            raise Exception(
                "Environment.__class__.default_logging_level can't be stringified. Check the value you entered while calling set_defaults() function."
            )
            
        logging_level = self._get_env(
            name="LOGGING_LEVEL", default=default
        )
        logging_level = logging_level.lower()

        levels = {
            "debug": logging.DEBUG,
            "info": logging.INFO,
            "warning": logging.WARNING,
            "error": logging.ERROR,
        }

        if logging_level in levels:
            return levels[logging_level]

        raise ValueError(
            f"Environment variable 'LOGGING_LEVEL' must be either 'debug', 'info', 'warning', or 'error'. Current value: '{logging_level}'."
        )
        
    def get_docker_image_version(self) -> str:
        """
        Return the Docker version from environment variable 'DOCKER_IMAGE'.
        """

        docker_version = self._get_env(name="DOCKER_IMAGE", default="unknown")
        
        if docker_version == "unknown":
            logging.warning("Couldn't get Docker version from env variable 'DOCKER_IMAGE'. (Ignore if you're running this script outside of Docker)")

        return docker_version

    def _get_env(
        self, name: str, default: str = "", strip: bool = True, info_log: bool = False
    ) -> str:
        """
        Get environment variable with optional default value.

        Logs a warning when default is used and an error when a required
        environment variable is missing.
        """

        env = os.getenv(name, "")

        if strip is True:
            env = env.strip()

        if env.strip() == "":  # if env only contains whitespace
            default = default.strip()
            if default:
                env = default

                logging_text = f"Environment Variable '{name}' is not set. Using default '{default}'."
                if info_log is True:
                    logging.info(logging_text)
                else:
                    logging.debug(logging_text)

            else:
                raise ValueError(f"Environment Variable '{name}' is not set.")

        env = self._decode_if_base64(env, prefix="BASE64:")

        return env

    def _get_secret(self, name: str) -> str:
        """
        Read a docker-style secret from a file and return its contents. Handles base64-encoded secrets if prefixed with "BASE64:".

        Returns an empty string and logs an error on failures.
        """

        path = os.path.join("/run/secrets", name)

        try:
            with open(path, "r") as f:
                secret = f.read().strip("\n")

        except Exception as e:
            logging.error(f"Error reading secret from {path}: {e}")
            secret = ""

        secret = self._decode_if_base64(secret, prefix="BASE64:")

        return secret

    def _decode_if_base64(self, data: str, prefix: str) -> str:
        """
        Decode the given data from base64 if it starts with the specified prefix.
        """

        if isinstance(data, str):
            stripped_data = data.strip()

            if stripped_data.startswith(prefix):
                try:
                    b64_content = stripped_data[len(prefix) :]
                    data = base64.b64decode(b64_content).decode("utf-8").strip("\n")
                except Exception as e:
                    logging.error(f"Error decoding base64: {e}")
                    data = ""

        return data


class PretixAPI:
    def __init__(self):
        """
        Initialize, fetch data from Pretix API.
        """
        env = Environment()

        pretix_url = env.get_pretix_url()
        pretix_organizer = env.get_pretix_orgnizer_slug()
        pretix_event = env.get_pretix_event_slug()
        pretix_api_token = env.get_pretix_api_token()

        self.pretix_api_url = os.path.join(
            pretix_url, "api/v1/organizers", pretix_organizer, "events", pretix_event
        )

        self.session = requests.Session()
        self.session.headers.update({"Authorization": f"Token {pretix_api_token}"})

        retries = Retry(
            total=5,
            backoff_factor=0.5,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["GET"],
        )

        adapter = HTTPAdapter(max_retries=retries)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)


    def _get_questions(self) -> dict:
        """
        Fetch all questions from Pretix API and return a mapping of question ID to question text.
        """

        url = f"{self.pretix_api_url}/questions/"
        questions = {}

        while url:
            r = self.session.get(url)
            r.raise_for_status()
            data = r.json()

            for q in data["results"]:
                question_text = q["question"].get("de") or next(
                    iter(q["question"].values())
                )
                questions[q["id"]] = question_text

            url = data["next"]  # Pagination

        return questions


    def get_question_choices_by_text(self, question_str: str) -> list:
        """
        Given the clear-text question name (question_str), find all question IDs that
        have that visible text and fetch their possible answer options from the
        Pretix API. If multiple question IDs match the same question_str, the
        returned answer options are merged uniquely and this fact is logged.

        Returns a sorted list of unique answer option strings.
        """

        if not question_str.strip():
            logging.error(
                "Empty question text provided to get_question_choices_by_text"
            )
            return []

        # build a fresh mapping of id -> text (ensures up-to-date data)
        question_map = self._get_questions()

        # case-insensitive match on the visible question text
        target = question_str.strip().lower()
        matching_qids = [
            qid
            for qid, text in question_map.items()
            if (text or "").strip().lower() == target
        ]

        if not matching_qids:
            logging.warning(f"No question IDs found for question text '{question_str}'")
            return []

        all_choices = set()

        for qid in matching_qids:
            url = f"{self.pretix_api_url}/questions/{qid}/"
            try:
                r = self.session.get(url)
                r.raise_for_status()
                data = r.json()

                # pretix may expose options under different keys depending on API version/implementation
                options = []
                for key in (
                    "options",
                    "choices",
                    "answers",
                    "options_list",
                    "question_options",
                ):
                    if key in data and isinstance(data[key], list):
                        options = data[key]
                        break

                # fallback: sometimes the question detail may include nested structures
                if not options:
                    # try to find any list-valued field in response that looks like options
                    for v in data.values():
                        if isinstance(v, list) and v and isinstance(v[0], (str, dict)):
                            options = v
                            break

                # extract a human-readable string from each option entry.
                def _extract_choice_text(opt) -> str | None:
                    # plain string option
                    if isinstance(opt, str):
                        return opt.strip() or None

                    if not isinstance(opt, dict):
                        return None

                    # prefer common direct string keys
                    for k in ("label", "text", "answer", "name", "title", "display"):
                        v = opt.get(k)
                        if isinstance(v, str) and v.strip():
                            return v.strip()

                        # if the value is a translations dict (e.g. {"de": "..."}), prefer German then English
                        if isinstance(v, dict):
                            for lang in ("de", "de-DE", "en", "en-US"):
                                if (
                                    lang in v
                                    and isinstance(v[lang], str)
                                    and v[lang].strip()
                                ):
                                    return v[lang].strip()
                            # fallback to first available string in translations
                            for vv in v.values():
                                if isinstance(vv, str) and vv.strip():
                                    return vv.strip()

                    # if no labelled text found, try to find any string value in the dict
                    for vv in opt.values():
                        if isinstance(vv, str) and vv.strip():
                            return vv.strip()

                    # nothing human-readable found
                    return None

                for opt in options:
                    text_val = _extract_choice_text(opt)
                    if text_val:
                        all_choices.add(text_val)

            except Exception as e:
                logging.error(f"Error fetching choices for question id {qid}: {e}")

        if len(matching_qids) > 1:
            logging.info(
                f"Multiple question IDs {matching_qids} map to the same question text '{question_str}'. Merged unique choices ({len(all_choices)} unique)."
            )

        return sorted(all_choices)


    def _get_items(self) -> dict:
        """
        Fetch all items from Pretix API and return a mapping of item ID to item name.
        """

        url = f"{self.pretix_api_url}/items/"
        items = {}

        while url:
            r = self.session.get(url)
            r.raise_for_status()
            data = r.json()

            for i in data["results"]:
                item_name = i["name"].get("de") or next(iter(i["name"].values()))
                items[i["id"]] = item_name

            url = data["next"]  # Pagination

        return items


    def _get_orders(self) -> list:
        """
        Fetch all orders from Pretix API and return as a list of order dicts.
        """

        url = f"{self.pretix_api_url}/orders/"
        orders = []

        while url:
            r = self.session.get(url)
            r.raise_for_status()
            data = r.json()

            orders.extend(data["results"])
            url = data["next"]  # Pagination

        return orders


    def _get_unique_column_name(self, base_name: str, existing_columns: list) -> str:
        """
        Generate a unique column name by adding (#2), (#3), etc. suffix if needed.
        """

        if base_name not in existing_columns:
            return base_name

        # Find the next available number
        counter = 2
        while f"{base_name} (#{counter})" in existing_columns:
            counter += 1

        return f"{base_name} (#{counter})"


    def get_raw_df(self) -> pd.DataFrame:
        """
        Fetch raw data from Pretix API and return as a pandas DataFrame.
        """

        question_map = self._get_questions()
        item_map = self._get_items()
        orders = self._get_orders()

        rows = []
        # Track which question texts we've already seen to handle duplicate question texts
        question_text_mapping = {}  # maps original qtext -> unique column name

        for order in orders:
            invoice = order.get("invoice_address", {}) or {}
            order_info = {
                "order_code": order["code"],
                "status": order["status"],
                "email": order["email"],
                "total": order["total"],
                "date": order["datetime"],
                "invoice_name": invoice.get("name", ""),
                "invoice_company": invoice.get("company", ""),
                "invoice_street": invoice.get("street", ""),
                "invoice_zipcode": invoice.get("zipcode", ""),
                "invoice_city": invoice.get("city", ""),
                "invoice_country": invoice.get("country", ""),
                "invoice_vat_id": invoice.get("vat_id", ""),
            }

            for position in order["positions"]:
                # resolve item id
                item_id = position.get("item")
                if isinstance(item_id, dict):
                    item_name = item_id["name"].get("de") or next(
                        iter(item_id["name"].values())
                    )
                else:
                    item_name = item_map.get(item_id, f"Item {item_id}")

                # attendee name from attendee_name_parts
                name_parts = position.get("attendee_name_parts", {}) or {}
                attendee_firstname = name_parts.get("given_name", "")
                attendee_lastname = name_parts.get("family_name", "")

                pos_info = {
                    "position_id": position["id"],
                    "item_id": item_id if not isinstance(item_id, dict) else item_id["id"],
                    "item_name": item_name,
                    "price": position["price"],
                    "attendee_firstname": attendee_firstname,
                    "attendee_lastname": attendee_lastname,
                }

                # answers for questions
                questions = {}
                for answer in position.get("answers", []):
                    qid = answer.get("question")
                    answer_text = answer.get("answer", "")

                    if isinstance(qid, dict):
                        qid = qid["id"]

                    qtext = question_map.get(qid, f"question_{qid}")

                    # Generate unique column name for this question text
                    if qtext not in question_text_mapping:
                        question_text_mapping[qtext] = self._get_unique_column_name(
                            qtext, list(question_text_mapping.values())
                        )

                    unique_col_name = question_text_mapping[qtext]
                    questions[unique_col_name] = answer_text

                row = {**order_info, **pos_info, **questions}
                rows.append(row)

        df = pd.DataFrame(rows)

        # Ensure all questions from Pretix are present as columns even if
        # no attendee has answered them yet. Handle duplicate question texts with (#2), (#3), etc.
        for qid, qtext in question_map.items():
            # Generate unique column name for this question text
            if qtext not in question_text_mapping:
                unique_col_name = self._get_unique_column_name(
                    qtext, list(question_text_mapping.values()) + list(df.columns)
                )
                question_text_mapping[qtext] = unique_col_name
            else:
                unique_col_name = question_text_mapping[qtext]

            # Add empty column if it doesn't exist yet
            if unique_col_name not in df.columns:
                df[unique_col_name] = ""

        logging.info("Fetched raw data from Pretix API.")

        return df


class Excel:
    def __init__(self):
        """
        Initialize the Excel helper class, setting up a temporary directory for Excel files.
        """
        env = Environment()
        self.max_column_width = env.get_excel_max_column_width()
        temp_dir_name = env.get_temp_dir_name()

        self.temp_dir = os.path.join(tempfile.gettempdir(), temp_dir_name)
        Path(self.temp_dir).mkdir(parents=True, exist_ok=True)
        

    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize the filename by replacing or removing invalid characters.
        """
        filename = filename.strip()
        filename = filename.replace("\\n", " ")
        filename = filename.replace("\\r", " ")
        filename = filename.replace("\\t", " ")
        filename = filename.replace("<", "_")
        filename = filename.replace(">", "_")
        filename = filename.replace(":", "")
        filename = filename.replace('"', "")
        filename = filename.replace("/", "+")
        filename = filename.replace("\\", "_")
        filename = filename.replace("|", "_")
        filename = filename.replace("?", "")
        filename = filename.replace("*", "")

        return filename

    def save_to_excel(self, df: pd.DataFrame, filename: str) -> str:
        """
        Save the given DataFrame to an Excel file with the specified filename in the temporary directory.
        Returns the path to the saved Excel file.
        """

        filename = self._sanitize_filename(filename)

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        sheet_name = filename.removesuffix(".xlsx")

        path = os.path.join(self.temp_dir, filename)

        # write dataframe to excel
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=True, freeze_panes=(1, 1))
            worksheet = writer.sheets[sheet_name]

            # adjust column widths including index column and header
            for idx, col in enumerate(
                df.columns, start=2
            ):  # start=2 because index is in column 1
                max_length = max(
                    len(str(col)),  # header length
                    df[col].astype(str).map(len).max(),  # data length
                )
                length = min(
                    max_length + 2, self.max_column_width
                )  # cap the width via constant

                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=idx).column_letter
                ].width = length

            # adjust index column width
            max_index_length = max(
                len(str(df.index.name or "")),  # index name length
                df.index.astype(str).map(len).max(),  # index data length
            )
            worksheet.column_dimensions["A"].width = max_index_length + 2

        logging.info(f"Created Excel file '{path}'.")

        return path

    def add_filters(self, path_to_excel_file: str) -> str:
        """
        Add filtering function to an existing Excel file.
        """

        path = path_to_excel_file

        if not os.path.isfile(path):
            logging.warning(f"Cannot add filter to '{path}': No such file.")
            return path

        elif not path.endswith(".xlsx"):
            logging.warning(
                f"Cannot add filter to '{path}': File is not an excel file."
            )
            return path

        wb = load_workbook(path)
        ws = wb.active

        # Determine the range of the data
        min_col = 2  # assuming index is in column A
        min_row = 1
        max_col = ws.max_column
        max_row = ws.max_row
        table_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

        # Create a table
        tab = Table(displayName="Table1", ref=table_range)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(tab)

        wb.save(path)

        logging.info(f"Added filter to Excel file '{path}'.")

        return path

    def delete_excel(self, path_to_excel_file: str) -> None:
        """
        Delete a temporary Excel file.
        """
        try:
            if os.path.isfile(path_to_excel_file):
                os.remove(path_to_excel_file)
                logging.info(f"Deleted temporary Excel file '{path_to_excel_file}'.")
        except Exception as e:
            logging.error(f"Error deleting file '{path_to_excel_file}': {e}")


class Nextcloud:
    def __init__(self):
        """
        Initialize Nextcloud connection with environment variables.
        """
        env = Environment()

        nextcloud_url = env.get_nextcloud_url()
        username = env.get_nextcloud_username()
        password = env.get_nextcloud_password()
        self.upload_dir = env.get_nextcloud_upload_dir()
        self.time_zone = env.get_timezone()

        self.base_url = os.path.join(
            nextcloud_url, "remote.php/dav/files", username
        )
        
        self.session = requests.Session()
        self.session.auth = HTTPBasicAuth(username, password)

        retries = Retry(
            total=5,
            backoff_factor=0.5,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["MKCOL", "PUT", "GET", "HEAD", "DELETE"]
        )

        adapter = HTTPAdapter(max_retries=retries)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        

    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize the filename by replacing or removing invalid characters.
        """
        filename = filename.strip()
        filename = filename.replace("\\n", " ")
        filename = filename.replace("\\r", " ")
        filename = filename.replace("\\t", " ")
        filename = filename.replace("<", "_")
        filename = filename.replace(">", "_")
        filename = filename.replace(":", "")
        filename = filename.replace('"', "")
        filename = filename.replace("/", "+")
        filename = filename.replace("\\", "_")
        filename = filename.replace("|", "_")
        filename = filename.replace("?", "")
        filename = filename.replace("*", "")

        return filename
    

    def _get_parent_directories(self, path: str) -> list[str]:
        """
        Get all parent directories of a given path.
        e.g. for path "A/B/C" it returns ["A", "A/B", "A/B/C"]
        """

        # ensure that the path does not end with a slash
        path = path.rstrip(os.sep)

        parent_dirs = []

        # walk through the path and build the parent directories
        while path != os.path.dirname(path):
            parent_dirs.append(path)
            path = os.path.dirname(path)

        # reverse the list to have the directories in hierarchical order from top to bottom.
        return parent_dirs[::-1]
            
        
    def create_dir(self, directory: str) -> None:
        """
        Create directory on Nextcloud.
        """
        
        full_dir = os.path.join(self.upload_dir, directory)
        webdav_url = os.path.join(self.base_url, full_dir)
        
        if "../" in full_dir or "/.." in full_dir:  # if directory tries to use parent directories and tries to upload to a destination outside of the given upload directory
            logging.warning(f"DO NOT USE '/../' segments in your directory path! This may create directories outside your upload directory! PROCEED WITH CAUTION ON YOUR OWN RISK!\nURL: {webdav_url}")
        
        try:
            # Try to create the direct directory first
            r = self.session.request(
                method="MKCOL",
                url=os.path.join(webdav_url),
            )

            if r.status_code == 405:
                logging.debug(f"Nextcloud directory already exists ({webdav_url})")
                return
            
            if r.status_code == 201:
                logging.debug(f"Created Nextcloud directory ({webdav_url})")
                logging.info(f"Created Nextcloud directory ({full_dir})")
                return
            
            if r.status_code == 409:
                logging.debug(f"Creating Nextcloud directory: Parent node does not exist ({webdav_url}). Creating parent directories now.")
                
                dir_path = os.path.join(self.upload_dir, directory)
                for dir in self._get_parent_directories(dir_path):
                    r = self.session.request(
                        method="MKCOL",
                        url=os.path.join(self.base_url, dir),
                    )

                    if r.status_code not in [201, 405]:
                        raise Exception(f"Error creating Nextcloud directory: {r.status_code} - {r.text}")

                logging.debug(f"Created Nextcloud directory ({webdav_url})")
                logging.info(f"Created Nextcloud directory ({full_dir})")

        except Exception as e:
            raise Exception(f"Error creating upload directory: {e}")
        
        
    
    def upload_file(self, filename: str, data: bytes, subdir: str = "") -> None:
        """
        Upload a file to Nextcloud via WebDAV. Destination of the uploaded file is the given Upload Directory in env variable NEXTCLOUD_UPLOAD_DIR plus optionally a given subdirectory.
        """
        
        # Add parent directories in filename to subdir (e.g. "A/B/file.txt" in subdir="A" and filename = "B/file.txt")
        filepath, filename = os.path.split(filename)
        subdir = os.path.join(subdir, filepath.strip("/\\"))
        
        upload_dir = os.path.join(self.upload_dir, subdir)
        
        if "../" in upload_dir or "/.." in upload_dir:  # if directory tries to use parent directories and tries to upload to a destination outside of the given upload directory
            raise Exception(f"DO NOT USE '/../' segments in your directory path! This may alter files outside your upload directory! PROCEED WITH CAUTION ON YOUR OWN RISK!\nURL: {upload_dir}")
        
        self.create_dir(subdir)            
        
        try:
            r = self.session.put(
                url=os.path.join(self.base_url, upload_dir, filename),
                data=data,
            )

            if r.status_code in (200, 201, 204):
                logging.info(f"File '{os.path.join(subdir, filename)}' uploaded successfully.")
            else:
                logging.error(
                    f"Error uploading file '{filename}': {r.status_code} - {r.text}"
                )

        except Exception as e:
            logging.error(f"Network error uploading file '{filename}': {e}")

    def upload_excel(self, source_file: str, subdir: str = "") -> None:
        """
        Upload an Excel file to Nextcloud.
        """

        try:
            filename = os.path.basename(source_file)
            filename = self._sanitize_filename(filename)
            if not filename.lower().endswith(".xlsx"):
                raise Exception("File is not an Excel file (.xlsx)")

            with open(source_file, "rb") as f:
                data = f.read()
            self.upload_file(filename, data, subdir)
        except Exception as e:
            logging.error(f"Error uploading Excel file '{source_file}': {e}")

    def upload_last_updated(self, filename = "Last_Updated.txt", subdir: str = "") -> None:
        """
        Upload a timestamp file indicating the last update time.
        """
        
        filename = self._sanitize_filename(filename)
        
        if not filename.lower().endswith(".txt"):
            filename += ".txt"

        tz = pytz.timezone(self.time_zone)
        data = "Last updated:\n" + datetime.now(tz=tz).strftime("%d.%m.%Y %H:%M")

        self.upload_file(filename, data.encode("utf-8"), subdir)
        
    def upload_docker_image_version(self, filename: str = "Docker_Image_Version.txt", subdir: str = "") -> None:
        """
        Upload a Docker image version file indicating the Docker image currently used.
        """
        
        filename = self._sanitize_filename(filename)
        
        if not filename.lower().endswith(".txt"):
            filename += ".txt"

        docker_image = Environment().get_docker_image_version()
        data = "Docker Image Base Version:\n" + docker_image

        self.upload_file(filename, data.encode("utf-8"), subdir)
        

class Main:
    def __init__(self):
        self.excel = Excel()
        self.nc = Nextcloud()

        self.success_on_last_run = False

        level = Environment().get_logging_level()

        # remove all handlers associated with the root logger object and set desired logging level
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        logging.basicConfig(level=level)


    def run(self):
        """
        Run main on startup and schedule loop for continuous execution.
        """
        
        # run main immediately on startup
        self.main_wrapper()

        # stop here if RUN_ONCE is true
        if Environment().get_run_once() is True:
            logging.info("RUN_ONCE is set to true. Exiting now after single run.")
            exit(0)

        # schedule loop for continuous execution
        self.schedule_loop()

    def upload(self, df: pd.DataFrame, filename: str, subdir: str = "", add_filters: bool = False):
        """
        Generate excel file from dataframe, upload excel file and delete it afterwards. Can also add filters to excel file.
        """
        
        filepath = self.excel.save_to_excel(df, filename)

        if add_filters is True:
            self.excel.add_filters(filepath)

        self.nc.upload_excel(filepath, subdir)

        self.excel.delete_excel(filepath)

    def schedule_loop(self):
        env = Environment()
        
        # schedule main to run at configured intervals
        schedule.every(env.get_interval_minutes()).minutes.do(self.main_wrapper)

        # keep the scheduler running
        while True:
            schedule.run_pending()
            time.sleep(
                env.get_check_interval_seconds()
            )  # check periodically if a task needs to run
            
    def upload_last_updated(self):
        try:
            self.nc.upload_last_updated()
        except Exception as e:
            raise Exception(f"An error occurred while uploading last updated timestamp: {e}")
        

    def main_wrapper(self):
        """
        Wrapper for main function with error handling.
        """

        try:
            self.excel = Excel()
            self.nc = Nextcloud()
            
            self.main()
            
            self.success_on_last_run = True

        except Exception as e:
            if str(e) == "No changes in data since last fetch.":
                logging.info("No changes in data detected since last fetch. Skipping upload process.")
                try:
                    self.upload_last_updated()
                except Exception as e:
                    logging.error(e)

            else:
                logging.error(f"An error occurred during execution: {e}")
                self.success_on_last_run = False

    def main(self):
        """
        You need to override this function using a subclass to customize your P2N instance.
        """
        
        error_str = "You need to override the main() function of the class Main using a subclass to customize your P2N instance !!!"
        logging.error(error_str)
        raise Exception(error_str)
